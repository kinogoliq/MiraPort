# gui.py

import sys
import os
import tempfile
import shutil
import subprocess
import logging
import tkinter as tk
from tkinter import messagebox, filedialog, E, W, N, S

import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from PIL import Image, ImageTk
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import Cell, MergedCell

from calculations import FeeCalculator
from constants import (
    START_ROW_FEES,
    START_ROW_AGENCY_FEES,
    LOGO_PATH,
    TEMPLATE_PATH,
    LOGO_PATH,
    TEMPLATE_PATH
)
from utils import format_amount, parse_input, resource_path
from agency_fee import calculate_cv, show_agency_fee_table, get_agency_fee
from fda_tab import FDATab

logger = logging.getLogger(__name__)


class ScrollableFrame(ttk.Frame):
    def __init__(self, container, *args, **kwargs):
        super().__init__(container, *args, **kwargs)

        self.canvas = tk.Canvas(self)
        scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = ttk.Frame(self.canvas)

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(
                scrollregion=self.canvas.bbox("all")
            )
        )

        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=scrollbar.set)

        self.canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Обработка прокрутки колесиком мыши
        self.bind_mousewheel(self.scrollable_frame)

    def bind_mousewheel(self, widget):
        if sys.platform.startswith('win'):
            widget.bind("<Enter>", self._bind_to_mousewheel_windows)
            widget.bind("<Leave>", self._unbind_from_mousewheel_windows)
        elif sys.platform.startswith('darwin'):
            widget.bind("<Enter>", self._bind_to_mousewheel_mac)
            widget.bind("<Leave>", self._unbind_from_mousewheel_mac)
        else:
            widget.bind("<Enter>", self._bind_to_mousewheel_linux)
            widget.bind("<Leave>", self._unbind_from_mousewheel_linux)
        for child in widget.winfo_children():
            self.bind_mousewheel(child)

    def _on_mousewheel_windows(self, event):
        self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    def _bind_to_mousewheel_windows(self, event):
        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel_windows)

    def _unbind_from_mousewheel_windows(self, event):
        self.canvas.unbind_all("<MouseWheel>")

    def _on_mousewheel_mac(self, event):
        self.canvas.yview_scroll(int(-1 * event.delta), "units")

    def _bind_to_mousewheel_mac(self, event):
        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel_mac)

    def _unbind_from_mousewheel_mac(self, event):
        self.canvas.unbind_all("<MouseWheel>")

    def _on_mousewheel_linux(self, event):
        if event.num == 4:
            self.canvas.yview_scroll(-1, "units")
        elif event.num == 5:
            self.canvas.yview_scroll(1, "units")

    def _bind_to_mousewheel_linux(self, event):
        self.canvas.bind_all("<Button-4>", self._on_mousewheel_linux)
        self.canvas.bind_all("<Button-5>", self._on_mousewheel_linux)

    def _unbind_from_mousewheel_linux(self, event):
        self.canvas.unbind_all("<Button-4>")
        self.canvas.unbind_all("<Button-5>")


class ProformaApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Расчет проформы дисбурсментского счета")
        self.root.geometry("1100x1000")  # Устанавливаем размер окна 800x600 пикселей
        # Установка иконки приложения
        self.set_app_icon()
        # Создаем стиль с выбранной темой
        self.style = ttk.Style(theme='cosmo')  # Вы можете выбрать другую тему
        self.pda_data = []  # Список fees и dues из PDA
        self.create_widgets()
        self.last_pdf_path = None  # Для хранения пути к последнему сгенерированному PDF
        self.cv = 0  # Инициализируем cv

    def set_app_icon(self):
        if sys.platform.startswith('win'):
            icon_path = resource_path(os.path.join('icons', 'app_icon.ico'))
            try:
                self.root.iconbitmap(icon_path)
            except Exception as e:
                logger.exception("Ошибка при установке иконки приложения: %s", e)
        elif sys.platform.startswith('darwin'):
            try:
                icon_path = resource_path(os.path.join('icons', 'app_icon.png'))
                image = Image.open(icon_path)
                photo = ImageTk.PhotoImage(image)
                self.root.iconphoto(False, photo)
            except Exception as e:
                logger.exception("Ошибка при установке иконки приложения: %s", e)
        else:
            icon_path = resource_path(os.path.join('icons', 'app_icon.png'))
            try:
                photo = ImageTk.PhotoImage(file=icon_path)
                self.root.iconphoto(False, photo)
            except Exception as e:
                logger.exception("Ошибка при установке иконки приложения: %s", e)
        pass

    def create_widgets(self):
        # Создаем Notebook
        notebook = ttk.Notebook(self.root)
        notebook.pack(fill=BOTH, expand=True)

        # Фреймы для вкладок
        self.input_frame = ttk.Frame(notebook)
        self.result_frame = ttk.Frame(notebook)
        self.fda_frame = ttk.Frame(notebook)  # Новая вкладка FDA

        notebook.add(self.input_frame, text='Ввод данных')
        notebook.add(self.result_frame, text='Результаты')
        notebook.add(self.fda_frame, text='FDA')  # Добавляем вкладку FDA

        # Вызов методов для создания виджетов
        self.create_input_widgets()
        self.create_result_widgets()

        # Создаем экземпляр вкладки FDA с пустыми данными
        self.fda_tab = FDATab(self.fda_frame, self.pda_data)

    def create_input_widgets(self):
        # Создаем прокручиваемый фрейм
        scrollable_frame = ScrollableFrame(self.input_frame)
        scrollable_frame.pack(fill="both", expand=True)

        # Теперь используем scrollable_frame.scrollable_frame для размещения виджетов
        container = scrollable_frame.scrollable_frame

        # Добавление логотип
        logo_frame = ttk.Frame(container)
        logo_frame.pack(pady=10)

        try:
            logo_image = Image.open(LOGO_PATH)
            logo_image = logo_image.resize((300, 300), Image.LANCZOS)
            self.logo_photo = ImageTk.PhotoImage(logo_image)
            logo_label = ttk.Label(logo_frame, image=self.logo_photo)
            logo_label.pack()
        except Exception as e:
            logger.exception("Необработанное исключение")
            print(f"Ошибка при загрузке логотипа: {e}")
            messagebox.showerror("Ошибка", f"Не удалось загрузить логотип: {e}")

        # Создание фрейма для полей ввода внутри прокручиваемого фрейма
        fields_frame = ttk.Frame(container)
        fields_frame.pack(pady=10, fill='x')

        # Поля ввода
        self.entries = {}

        # Левый и правый фреймы
        left_frame = ttk.Frame(fields_frame)
        left_frame.pack(side=LEFT, padx=10, fill=BOTH, expand=True)

        right_frame = ttk.Frame(fields_frame)
        right_frame.pack(side=LEFT, padx=10, fill=BOTH, expand=True)

        # Левые поля ввода
        left_fields = [
            ("LBP:", 'lbp'),
            ("Beam:", 'beam'),
            ("RDM:", 'rdm'),
            ("Количество миль внутренней проводки (In):", 'miles_inward_in'),
            ("Количество миль внутренней проводки (Out):", 'miles_inward_out'),
            ("Количество миль внешней проводки (In):", 'miles_outward_in'),
            ("Количество миль внешней проводки (Out):", 'miles_outward_out'),
            ("Agency fee:", 'agency_fee'),
            ("Bank charges:", 'bank_charges'),
        ]

        for label_text, var_name in left_fields:
            frame = ttk.Frame(left_frame)
            frame.pack(fill=X, pady=5)
            label = ttk.Label(frame, text=label_text, width=35, anchor=E)
            label.pack(side=LEFT)
            entry = ttk.Entry(frame)
            entry.pack(side=LEFT, fill=X, expand=True)
            self.entries[var_name] = entry

        # Правые поля ввода
        right_fields = [
            ("Vessel name:", 'vessel_name'),
            ("Vessel flag:", 'vessel_flag'),
            ("Enter Port:", 'enter_port'),
            ("Cargo loaded:", 'cargo_loaded'),
            ("Quantity of cargo, mts:", 'cargo_qtty'),
            ("Acc name:", 'acc_name'),
            ("VAT (%):", 'vat'),
            ("Overtime in:", 'overtime_in'),
            ("Overtime out:", 'overtime_out'),
        ]

        for label_text, var_name in right_fields:
            frame = ttk.Frame(right_frame)
            frame.pack(fill=X, pady=5)
            label = ttk.Label(frame, text=label_text, width=25, anchor=E)
            label.pack(side=LEFT)
            if var_name in ['overtime_in', 'overtime_out']:
                combobox = ttk.Combobox(frame, values=["0%", "25%", "50%", "100%"], state="readonly")
                combobox.set("0%")
                combobox.pack(side=LEFT, fill=X, expand=True)
                self.entries[var_name] = combobox
            else:
                entry = ttk.Entry(frame)
                entry.pack(side=LEFT, fill=X, expand=True)
                self.entries[var_name] = entry

        # Значения по умолчанию
        self.entries['miles_inward_in'].insert(0, "1")
        self.entries['miles_inward_out'].insert(0, "1")
        self.entries['miles_outward_in'].insert(0, "14")
        self.entries['miles_outward_out'].insert(0, "14")
        self.entries['agency_fee'].insert(0, "0")
        self.entries['bank_charges'].insert(0, "190.00")
        self.entries['vat'].insert(0, "20")

        # Кнопка расчета CV
        calculate_cv_button = ttk.Button(
            container,
            text="Рассчитать CV и Agency Fee",
            command=self.calculate_cv_and_agency_fee,
            bootstyle='primary'
        )
        calculate_cv_button.pack(pady=10)

        # Кнопка расчета с иконкой
        # Загрузка иконки
        try:
            calculate_icon = Image.open('icons/calculate_icon.png')
            calculate_icon = calculate_icon.resize((24, 24), Image.LANCZOS)
            calculate_icon_photo = ImageTk.PhotoImage(calculate_icon)
        except Exception as e:
            logger.exception("Необработанное исключение")
            print(f"Ошибка при загрузке иконки: {e}")
            calculate_icon_photo = None

        calculate_button = ttk.Button(
            container,
            text=" Рассчитать",
            image=calculate_icon_photo,
            compound=LEFT,
            command=self.calculate,
            bootstyle='primary'
        )
        if calculate_icon_photo:
            calculate_button.image = calculate_icon_photo  # Сохранение ссылки на изображение
        calculate_button.pack(pady=20)

        # Фрейм для дополнительных Dues
        dues_frame = ttk.Labelframe(container, text="Дополнительные Dues")
        dues_frame.pack(fill=X, padx=10, pady=10)

        self.additional_dues = []

        def add_additional_due():
            due_frame = ttk.Frame(dues_frame)
            due_frame.pack(fill=X, pady=2)

            due_name_entry = ttk.Entry(due_frame)
            due_name_entry.pack(side=LEFT, padx=5, pady=5, fill=X, expand=True)
            due_name_entry.insert(0, "Название Due")

            due_amount_entry = ttk.Entry(due_frame, width=15)
            due_amount_entry.pack(side=LEFT, padx=5, pady=5)
            due_amount_entry.insert(0, "Сумма")

            remove_button = ttk.Button(due_frame, text="Удалить", command=lambda: self.remove_additional_due(due_frame))
            remove_button.pack(side=LEFT, padx=5, pady=5)

            self.additional_dues.append((due_name_entry, due_amount_entry))

        add_due_button = ttk.Button(dues_frame, text="Добавить Due", command=add_additional_due, bootstyle='success')
        add_due_button.pack(anchor='w', padx=5, pady=5)

        # Фрейм для дополнительных Fees
        fees_frame = ttk.Labelframe(container, text="Дополнительные Fees")
        fees_frame.pack(fill=X, padx=10, pady=10)

        self.additional_fees = []

        def add_additional_fee():
            fee_frame = ttk.Frame(fees_frame)
            fee_frame.pack(fill=X, pady=2)

            fee_name_entry = ttk.Entry(fee_frame)
            fee_name_entry.pack(side=LEFT, padx=5, pady=5, fill=X, expand=True)
            fee_name_entry.insert(0, "Название Fee")

            fee_amount_entry = ttk.Entry(fee_frame, width=15)
            fee_amount_entry.pack(side=LEFT, padx=5, pady=5)
            fee_amount_entry.insert(0, "Сумма")

            remove_button = ttk.Button(fee_frame, text="Удалить", command=lambda: self.remove_additional_fee(fee_frame))
            remove_button.pack(side=LEFT, padx=5, pady=5)

            self.additional_fees.append((fee_name_entry, fee_amount_entry))

        add_fee_button = ttk.Button(fees_frame, text="Добавить Fee", command=add_additional_fee, bootstyle='success')
        add_fee_button.pack(anchor='w', padx=5, pady=5)

    def calculate_cv_and_agency_fee(self):
        """Метод для расчёта CV и Agency Fee при нажатии кнопки."""
        try:
            lbp_str = self.entries['lbp'].get().strip()
            beam_str = self.entries['beam'].get().strip()
            rdm_str = self.entries['rdm'].get().strip()

            if not lbp_str or not beam_str or not rdm_str:
                messagebox.showwarning("Предупреждение", "Пожалуйста, заполните все поля LBP, Beam и RDM.")
                return

            lbp = parse_input(lbp_str)
            beam = parse_input(beam_str)
            rdm = parse_input(rdm_str)

            self.cv = calculate_cv(lbp, beam, rdm)
            logger.info(f"CV рассчитан: {self.cv}")

            # Обновляем поле Agency Fee с полученным значением
            agency_fee = get_agency_fee(self.cv)
            if agency_fee is not None:
                self.entries['agency_fee'].delete(0, tk.END)
                self.entries['agency_fee'].insert(0, format_amount(agency_fee))
            else:
                messagebox.showwarning("Внимание", "CV не соответствует ни одному диапазону в таблице Agency Fee.")

            # Отображаем всплывающее окно с таблицей agency fee
            show_agency_fee_table(self.cv)

        except ValueError as e:
            logger.error(f"Ошибка при вводе размеров: {e}")
            messagebox.showerror("Ошибка ввода",
                                 "Пожалуйста, введите корректные числовые значения для LBP, Beam и RDM.")

    def validate_numeric_input(self, action, value_if_allowed):
        if action != '1':  # Если не вставка символа, пропускаем
            return True
        try:
            float(value_if_allowed.replace(',', '.'))
            return True
        except ValueError:
            return False

    def remove_additional_due(self, due_frame):
        due_frame.destroy()
        self.additional_dues = [due for due in self.additional_dues if due[0].winfo_exists()]

    def remove_additional_fee(self, fee_frame):
        fee_frame.destroy()
        self.additional_fees = [fee for fee in self.additional_fees if fee[0].winfo_exists()]

    def create_result_widgets(self):
        # Информационные метки
        info_frame = ttk.Frame(self.result_frame)
        info_frame.pack(pady=10, fill='x')

        self.port_label = ttk.Label(info_frame, text="Enter Port: ")
        self.port_label.pack(anchor='w')
        self.vessel_name_label = ttk.Label(info_frame, text="Vessel name: ")
        self.vessel_name_label.pack(anchor='w')
        self.vessel_flag_label = ttk.Label(info_frame, text="Vessel flag: ")
        self.vessel_flag_label.pack(anchor='w')
        self.cargo_loaded_label = ttk.Label(info_frame, text="Cargo loaded: ")
        self.cargo_loaded_label.pack(anchor='w')
        self.cargo_qtty_label = ttk.Label(info_frame, text="Quantity of cargo, mts: ")
        self.cargo_qtty_label.pack(anchor='w')
        self.acc_name_label = ttk.Label(info_frame, text="Acc name: ")
        self.acc_name_label.pack(anchor='w')

        # Добавление меток для "Agency fee" и "Bank charges"
        self.agency_fee_label = ttk.Label(info_frame, text="Agency fee: ")
        self.agency_fee_label.pack(anchor='w')
        self.bank_charges_label = ttk.Label(info_frame, text="Bank charges: ")
        self.bank_charges_label.pack(anchor='w')

        # Таблица результатов
        table_frame = ttk.Frame(self.result_frame)
        table_frame.pack(fill=BOTH, expand=True, pady=10)

        columns = ("Fee", "VAT", "Amount")
        self.tree = ttk.Treeview(table_frame, columns=columns, show="headings")
        self.tree.heading("Fee", text="Наименование сбора")
        self.tree.heading("VAT", text="VAT")
        self.tree.heading("Amount", text="Сумма")

        self.tree.column("Fee", width=300, anchor="w")
        self.tree.column("VAT", width=100, anchor="e")
        self.tree.column("Amount", width=100, anchor="e")

        self.tree.pack(side='left', fill='both', expand=True)

        scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side='right', fill='y')

        # Итоговые суммы
        totals_frame = ttk.Frame(self.result_frame)
        totals_frame.pack(pady=10, fill='x')

        self.subtotal_dues_label = ttk.Label(totals_frame, text="Subtotal (Dues): ")
        self.subtotal_dues_label.pack(anchor='w')
        self.subtotal_agfee_label = ttk.Label(totals_frame, text="Subtotal Agency Fees: ")
        self.subtotal_agfee_label.pack(anchor='w')
        self.total_label = ttk.Label(totals_frame, text="Total: ")
        self.total_label.pack(anchor='w')

        # Кнопки действий
        action_frame = ttk.Frame(self.result_frame)
        action_frame.pack(pady=10)

        save_icon = None
        print_icon = None
        display_icon = None

        try:
            save_image = Image.open('icons/save_icon.png')
            save_image = save_image.resize((24, 24), Image.LANCZOS)
            save_icon = ImageTk.PhotoImage(save_image)

            print_image = Image.open('icons/print_icon.png')
            print_image = print_image.resize((24, 24), Image.LANCZOS)
            print_icon = ImageTk.PhotoImage(print_image)

            display_image = Image.open('icons/display_icon.png')
            display_image = display_image.resize((24, 24), Image.LANCZOS)
            display_icon = ImageTk.PhotoImage(display_image)
        except Exception as e:
            logger.exception("Необработанное исключение")
            print(f"Ошибка при загрузке иконок: {e}")

        save_pdf_button = ttk.Button(
            action_frame,
            text=" Сохранить в PDF",
            image=save_icon,
            compound=LEFT,
            command=self.save_pdf,
            bootstyle='primary'
        )
        if save_icon:
            save_pdf_button.image = save_icon
        save_pdf_button.pack(side=LEFT, padx=5)

        print_button = ttk.Button(
            action_frame,
            text=" Печать",
            image=print_icon,
            compound=LEFT,
            command=self.print_result,
            bootstyle='primary'
        )
        if print_icon:
            print_button.image = print_icon
        print_button.pack(side=LEFT, padx=5)

        display_button = ttk.Button(
            action_frame,
            text=" Вывести на экран",
            image=display_icon,
            compound=LEFT,
            command=self.display_pdf,
            bootstyle='primary'
        )
        if display_icon:
            display_button.image = display_icon
        display_button.pack(side=LEFT, padx=5)

        # Создаём фрейм для итогов по фиксированным ставкам овертайма
        fixed_totals_frame = ttk.Frame(self.result_frame)
        fixed_totals_frame.pack(fill=BOTH, expand=True, pady=10)

        # Добавляем заголовок
        fixed_totals_label = ttk.Label(fixed_totals_frame, text="Итоги по фиксированным ставкам овертайма",
                                       font=('Helvetica', 12, 'bold'))
        fixed_totals_label.pack(anchor='w')

        # Создаём таблицу для итогов
        fixed_totals_columns = ("Description", "Amount")
        self.fixed_totals_tree = ttk.Treeview(fixed_totals_frame, columns=fixed_totals_columns, show="headings")
        self.fixed_totals_tree.heading("Description", text="Описание")
        self.fixed_totals_tree.heading("Amount", text="Сумма")

        self.fixed_totals_tree.column("Description", width=300, anchor="w")
        self.fixed_totals_tree.column("Amount", width=100, anchor="e")

        self.fixed_totals_tree.pack(side='left', fill='both', expand=True)

        fixed_totals_scrollbar = ttk.Scrollbar(fixed_totals_frame, orient="vertical",
                                               command=self.fixed_totals_tree.yview)
        self.fixed_totals_tree.configure(yscrollcommand=fixed_totals_scrollbar.set)
        fixed_totals_scrollbar.pack(side='right', fill='y')

    def get_input_values(self):
        inputs = {}
        for key, entry in self.entries.items():
            inputs[key] = entry.get()
        return inputs

    def calculate(self):
        logger.info("Начало расчета")
        inputs = self.get_input_values()
        # Собираем дополнительные Dues и Fees
        inputs['additional_dues'] = []
        for due_name_entry, due_amount_entry in self.additional_dues:
            due_name = due_name_entry.get()
            due_amount = due_amount_entry.get()
            inputs['additional_dues'].append({'name': due_name, 'amount': due_amount})

        inputs['additional_fees'] = []
        for fee_name_entry, fee_amount_entry in self.additional_fees:
            fee_name = fee_name_entry.get()
            fee_amount = fee_amount_entry.get()
            inputs['additional_fees'].append({'name': fee_name, 'amount': fee_amount})

        try:
            self.calculator = FeeCalculator(inputs)
            self.calculator.calculate_fees()
            self.calculator.calculate_totals()
            # Вызов нового метода для расчёта фиксированных ставок овертайма
            self.calculator.calculate_fixed_overtime_totals()
            self.update_results()
            logger.info("Расчет успешно завершен")
        except Exception as e:
            logger.error(f"Ошибка при расчете: {e}")
            messagebox.showerror("Ошибка", str(e))

        # После успешного расчёта сохраняем данные PDA
        self.pda_data = self.calculator.get_fees_and_dues()
        logger.info(f"PDA Data: {self.pda_data}")  # Добавляем логирование для проверки данных
        # Обновляем вкладку FDA с новыми данными
        if hasattr(self, 'fda_tab'):
            self.fda_tab.update_pda_data(self.pda_data)
            logger.info("FDA tab updated with new PDA data")

    def update_results(self):
        # Очистка предыдущих результатов
        for item in self.tree.get_children():
            self.tree.delete(item)

        # Обновление информационных меток
        self.port_label.config(text=f"Enter Port: {self.entries['enter_port'].get()}")
        self.vessel_name_label.config(text=f"Vessel name: {self.entries['vessel_name'].get()}")
        self.vessel_flag_label.config(text=f"Vessel flag: {self.entries['vessel_flag'].get()}")
        self.cargo_loaded_label.config(text=f"Cargo loaded: {self.entries['cargo_loaded'].get()}")
        self.cargo_qtty_label.config(text=f"Quantity of cargo, mts: {self.entries['cargo_qtty'].get()}")
        self.acc_name_label.config(text=f"Acc name: {self.entries['acc_name'].get()}")

        # Обновление меток "Agency fee" и "Bank charges"
        try:
            agency_fee_value = parse_input(self.entries['agency_fee'].get())
            bank_charges_value = parse_input(self.entries['bank_charges'].get())
            self.agency_fee_label.config(text=f"Agency fee: {format_amount(agency_fee_value)}")
            self.bank_charges_label.config(text=f"Bank charges: {format_amount(bank_charges_value)}")
        except ValueError as e:
            messagebox.showerror("Ошибка ввода", str(e))
            return

        # Заполнение таблицы результатов (только Dues)
        dues_data = self.calculator.get_fee_display_data()
        for fee_data in dues_data:
            if fee_data[0] not in ["Agency fee", "Bank charges"]:
                self.tree.insert("", "end", values=fee_data)

        # Добавление пустой строки для разделения
        self.tree.insert("", "end", values=("", "", ""))

        # Очистка предыдущих результатов в fixed_totals_tree
        for item in self.fixed_totals_tree.get_children():
            self.fixed_totals_tree.delete(item)

        # Добавление итогов по фиксированным ставкам овертайма
        for rate in sorted(self.calculator.fixed_totals.keys()):
            totals = self.calculator.fixed_totals[rate]
            percentage = int(rate * 100)
            self.fixed_totals_tree.insert("", "end", values=(
                f"Total fee with {percentage}% overtime", format_amount(totals['total_fee'])))
            self.fixed_totals_tree.insert("", "end", values=(
                f"Total agency fee (Basis {percentage}% overtime)", format_amount(totals['total_agency_fee'])))
            self.fixed_totals_tree.insert("", "end", values=(
                f"Grand total basis {percentage}% overtime", format_amount(totals['grand_total'])))
            # Добавляем пустую строку для разделения
            self.fixed_totals_tree.insert("", "end", values=("", ""))

        # Обновление итоговых сумм
        self.subtotal_dues_label.config(text=f"Subtotal (Dues): {format_amount(self.calculator.subtotal_dues)}")
        self.subtotal_agfee_label.config(
            text=f"Subtotal Agency Fees: {format_amount(self.calculator.subtotal_agency_fees)}")
        self.total_label.config(text=f"Total: {format_amount(self.calculator.total_amount)}")

    def save_pdf(self):
        if not hasattr(self, 'calculator'):
            messagebox.showwarning("Предупреждение", "Сначала выполните расчет.")
            return

        file_path = filedialog.asksaveasfilename(defaultextension=".pdf",
                                                 filetypes=[("PDF files", "*.pdf")])
        if not file_path:
            return

        try:
            self.generate_pdf(file_path)
            messagebox.showinfo("Успех", "Файл успешно сохранен.")
        except Exception as e:
            logger.exception("Необработанное исключение")
            messagebox.showerror("Ошибка", f"Не удалось сохранить файл: {e}")

    def print_result(self):
        if not hasattr(self, 'calculator'):
            messagebox.showwarning("Предупреждение", "Сначала выполните расчет.")
            return

        try:
            tmp_dir = tempfile.gettempdir()
            tmp_pdf_path = os.path.join(tmp_dir, next(tempfile._get_candidate_names()) + '.pdf')
            self.generate_pdf(tmp_pdf_path)

            if sys.platform.startswith('win'):
                os.startfile(tmp_pdf_path, "print")
            elif sys.platform.startswith('darwin') or sys.platform.startswith('linux'):
                subprocess.run(['lp', tmp_pdf_path])
            else:
                messagebox.showwarning("Предупреждение",
                                       "Неизвестная операционная система. Не удаётся отправить на печать автоматически.")
                return

            os.remove(tmp_pdf_path)
            messagebox.showinfo("Успех", "Документ успешно отправлен на печать.")
        except Exception as e:
            logger.exception("Необработанное исключение")
            messagebox.showerror("Ошибка", f"Не удалось напечатать файл: {e}")

    def display_pdf(self):
        if not hasattr(self, 'calculator'):
            messagebox.showwarning("Предупреждение", "Сначала выполните расчет.")
            return

        try:
            tmp_dir = tempfile.gettempdir()
            tmp_pdf_path = os.path.join(tmp_dir, next(tempfile._get_candidate_names()) + '.pdf')
            self.generate_pdf(tmp_pdf_path)

            if sys.platform.startswith('darwin'):
                subprocess.run(['open', tmp_pdf_path])
            elif sys.platform.startswith('win'):
                os.startfile(tmp_pdf_path)
            elif sys.platform.startswith('linux'):
                subprocess.run(['xdg-open', tmp_pdf_path])
            else:
                messagebox.showwarning("Предупреждение",
                                       "Неизвестная операционная система. Не удаётся открыть PDF автоматически.")
                return

            self.last_pdf_path = tmp_pdf_path

        except Exception as e:
            logger.exception("Необработанное исключение")
            messagebox.showerror("Ошибка", f"Не удалось открыть файл: {e}")

    def generate_pdf(self, pdf_path):
        logger.info(f"Начало генерации PDF по пути: {pdf_path}")
        # Проверка наличия шаблона Excel
        if not os.path.exists(TEMPLATE_PATH):
            messagebox.showerror("Ошибка",
                                 f"Шаблон Excel не найден. Убедитесь, что '{TEMPLATE_PATH}' находится в директории проекта.")
            return

        # Загрузка шаблона Excel
        try:
            logger.debug("Загрузка шаблона Excel")
            wb = openpyxl.load_workbook(TEMPLATE_PATH)
            ws = wb.active
        except Exception as e:
            logger.error(f"Ошибка при загрузке шаблона Excel: {e}")
            messagebox.showerror("Ошибка", f"Не удалось загрузить шаблон Excel: {e}")
            return

        # Подготовка данных для замены
        replacements = {
            '{{cv}}': format_amount(self.calculator.cv),
            '{{enter_port}}': self.entries['enter_port'].get(),
            '{{vessel_name}}': self.entries['vessel_name'].get(),
            '{{vessel_flag}}': self.entries['vessel_flag'].get(),
            '{{cargo_loaded}}': self.entries['cargo_loaded'].get(),
            '{{cargo_qtty}}': self.entries['cargo_qtty'].get(),
            '{{lbp}}': format_amount(parse_input(self.entries['lbp'].get())),
            '{{beam}}': format_amount(parse_input(self.entries['beam'].get())),
            '{{rdm}}': format_amount(parse_input(self.entries['rdm'].get())),
            '{{Account_name}}': self.entries['acc_name'].get(),
            '{{subtotal_dues}}': format_amount(self.calculator.subtotal_dues),
            '{{subtotal_agfee}}': format_amount(self.calculator.subtotal_agency_fees),
            '{{total}}': format_amount(self.calculator.total_amount),
            '{{total_vat}}': format_amount(self.calculator.total_vat),
            # Добавление Agency fee и Bank charges
            '{{agency_fee}}': format_amount(parse_input(self.entries['agency_fee'].get())),
            '{{bank_charges}}': format_amount(parse_input(self.entries['bank_charges'].get())),
            # Добавление новых плейсхолдеров для фиксированных ставок овертайма
            '{{total_fee_25_ot}}': format_amount(self.calculator.fixed_totals[0.25]['total_fee']),
            '{{total_agency_fee_25_ot}}': format_amount(self.calculator.fixed_totals[0.25]['total_agency_fee']),
            '{{grand_total_25_ot}}': format_amount(self.calculator.fixed_totals[0.25]['grand_total']),
            '{{total_fee_50_ot}}': format_amount(self.calculator.fixed_totals[0.50]['total_fee']),
            '{{total_agency_fee_50_ot}}': format_amount(self.calculator.fixed_totals[0.50]['total_agency_fee']),
            '{{grand_total_50_ot}}': format_amount(self.calculator.fixed_totals[0.50]['grand_total']),
            '{{total_fee_100_ot}}': format_amount(self.calculator.fixed_totals[1.00]['total_fee']),
            '{{total_agency_fee_100_ot}}': format_amount(self.calculator.fixed_totals[1.00]['total_agency_fee']),
            '{{grand_total_100_ot}}': format_amount(self.calculator.fixed_totals[1.00]['grand_total']),
        }

        # Замена плейсхолдеров в шаблоне
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    for key, value in replacements.items():
                        if key in cell.value:
                            cell.value = cell.value.replace(key, value)

        # Добавление таблицы сборов
        start_row = START_ROW_FEES  # Начальная строка для вывода сборов
        current_row = start_row

        # Заполнение таблицы сборов (только Dues)
        for fee in self.calculator.fees:
            if fee.name not in ["Agency fee", "Bank charges"]:
                ws.cell(row=current_row, column=1).value = fee.name
                ws.cell(row=current_row, column=5).value = format_amount(fee.vat_amount) if fee.vat_amount > 0 else "-"
                ws.cell(row=current_row, column=7).value = format_amount(fee.total_amount)
                current_row += 1

        # Устанавливаем фиксированную строку для Agency fee и Bank charges
        agency_start_row = START_ROW_AGENCY_FEES  # Укажите нужный номер строки
        ws.cell(row=agency_start_row, column=1).value = "Agency fee"
        ws.cell(row=agency_start_row, column=7).value = format_amount(parse_input(self.entries['agency_fee'].get()))
        agency_start_row += 1

        ws.cell(row=agency_start_row, column=1).value = "Bank charges"
        ws.cell(row=agency_start_row, column=7).value = format_amount(parse_input(self.entries['bank_charges'].get()))
        agency_start_row += 1

        # Добавление дополнительных Fees после Agency fee и Bank charges
        for fee in self.calculator.additional_fees:
            ws.cell(row=agency_start_row, column=1).value = fee['name']
            ws.cell(row=agency_start_row, column=7).value = format_amount(fee['amount'])
            agency_start_row += 1

        # Сохранение заполненного Excel во временный файл
        tmp_dir = tempfile.gettempdir()
        tmp_name = next(tempfile._get_candidate_names()) + '.xlsx'
        tmp_path = os.path.join(tmp_dir, tmp_name)
        wb.save(tmp_path)

        # Конвертация Excel в PDF
        soffice_path = self.get_soffice_path()
        if not soffice_path:
            return

        conversion_command = [
            soffice_path,
            '--headless',
            '--convert-to',
            'pdf',
            '--outdir',
            tmp_dir,
            tmp_path
        ]

        conversion_result = subprocess.run(conversion_command, capture_output=True, text=True)
        if conversion_result.returncode != 0:
            messagebox.showerror("Ошибка", f"Ошибка при конвертации Excel в PDF:\n{conversion_result.stderr}")
            return

        # Получение пути к сгенерированному PDF
        base_name = os.path.splitext(tmp_name)[0]
        pdf_tmp_name = base_name + '.pdf'
        pdf_tmp_path = os.path.join(tmp_dir, pdf_tmp_name)

        if not os.path.exists(pdf_tmp_path):
            messagebox.showerror("Ошибка", "Сгенерированный PDF-файл не найден.")
            return

        shutil.copy(pdf_tmp_path, pdf_path)

        # Удаление временных файлов
        os.remove(tmp_path)
        os.remove(pdf_tmp_path)

    def get_soffice_path(self):
        soffice_path = ""
        if sys.platform.startswith('darwin'):
            soffice_path = "/Applications/LibreOffice.app/Contents/MacOS/soffice"
        elif sys.platform.startswith('win'):
            soffice_path = os.path.join(os.environ.get("PROGRAMFILES", "C:\\Program Files"), "LibreOffice", "program",
                                        "soffice.exe")
            if not os.path.exists(soffice_path):
                soffice_path = os.path.join(os.environ.get("PROGRAMFILES(X86)", "C:\\Program Files (x86)"),
                                            "LibreOffice", "program", "soffice.exe")
        elif sys.platform.startswith('linux'):
            soffice_path = "/usr/bin/soffice"
        else:
            messagebox.showerror("Ошибка",
                                 "Неизвестная операционная система. Необходимо вручную указать путь к 'soffice'.")
            return None

        if not os.path.exists(soffice_path):
            messagebox.showerror("Ошибка", f"Не удалось найти soffice по пути: {soffice_path}")
            return None

        return soffice_path


if __name__ == "__main__":
    root = tk.Tk()
    app = ProformaApp(root)
    root.mainloop()
